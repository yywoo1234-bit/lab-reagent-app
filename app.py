import streamlit as st
import pandas as pd
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =================================================
# 기본 설정
# =================================================
st.set_page_config(
    page_title="🧪 시약 유통기한 자동 관리",
    layout="wide"
)

FILE_NAME = "reagents.xlsx"

# =================================================
# 데이터 로드
# =================================================
@st.cache_data
def load_data():
    # [수정] 터미널 로그 출력: 데이터 로딩 시작 알림
    print(f"\n[System Log] '{FILE_NAME}' 파일을 불러오는 중입니다...")
    
    try:
        df = pd.read_excel(FILE_NAME)
        df['등록일'] = pd.to_datetime(df['등록일'], errors="coerce")
        df['유통기한'] = pd.to_datetime(df['유통기한'], errors="coerce")
        
        # [수정] 데이터 로드 성공 로그
        print(f"[System Log] 데이터 로드 성공! 총 {len(df)}개의 시약 데이터가 있습니다.")
        return df
    except Exception as e:
        print(f"[Error] 파일을 찾을 수 없거나 읽는 중 오류 발생: {e}")
        return pd.DataFrame() # 빈 데이터프레임 반환

df = load_data()

# =================================================
# 날짜 계산
# =================================================
today = pd.to_datetime(datetime.today().date())
df['남은일수'] = (df['유통기한'] - today).dt.days
df = df.sort_values(by='남은일수')

# 분류
expired = df[df['남은일수'] < 0]
soon = df[(df['남은일수'] >= 0) & (df['남은일수'] <= 30)]
safe = df[df['남은일수'] > 30]

# [수정] 터미널에 현재 상태 요약 리포트 출력 (교수님께 보여드리기 좋은 부분)
print("-" * 30)
print(f"기준일: {today.date()}")
print(f"🔴 폐기 대상: {len(expired)}건")
print(f"🟡 임박 시약: {len(soon)}건")
print(f"⚪ 안전 시약: {len(safe)}건")
print("-" * 30)

# =================================================
# 화면 표시
# =================================================
st.title("🧪 시약 유통기한 자동 관리 시스템")
st.write(f"📅 기준일: **{today.date()}**")

def color_df(row):
    if row['남은일수'] < 0:
        return ['background-color:#ffcccc'] * len(row)
    elif row['남은일수'] <= 30:
        return ['background-color:#fff2cc'] * len(row)
    return ['background-color:white'] * len(row)

# =================================================
# 🚨 1. 유통기한 지난 시약
# =================================================
st.subheader("🔴 유통기한 지난 시약")

if expired.empty:
    st.success("✅ 유통기한이 지난 시약이 없습니다.")
else:
    # [수정] 데이터가 있을 때만 터미널에 경고 메시지 출력
    print(f"[Warning] 폐기해야 할 시약이 {len(expired)}개 발견되었습니다.")
    st.dataframe(expired.style.apply(color_df, axis=1), use_container_width=True)

# =================================================
# ⚠️ 2. 유통기한 임박 시약
# =================================================
st.subheader("🟡 유통기한 임박 시약 (30일 이내)")

if soon.empty:
    st.success("✅ 유통기한 임박 시약이 없습니다.")
else:
    st.dataframe(soon.style.apply(color_df, axis=1), use_container_width=True)

# =================================================
# ✅ 3. 유통기한 충분히 남은 시약
# =================================================
st.subheader("⚪ 유통기한 충분히 남은 시약")

if safe.empty:
    st.info("표시할 시약이 없습니다.")
else:
    st.dataframe(safe.style.apply(color_df, axis=1), use_container_width=True)

# =================================================
# 🔍 4. 전체 시약 통합 검색
# =================================================
st.divider()
st.subheader("🔍 전체 시약 검색")

search_term = st.text_input("시약 제품명 입력 (부분 검색 가능)")

search_df = df.copy()

if search_term:
    # [수정] 검색 기능 사용 시 터미널에 검색어 기록
    print(f"[User Action] 사용자가 '{search_term}'을(를) 검색했습니다.")
    
    search_df = search_df[
        search_df['제품명'].astype(str).str.contains(search_term, case=False, na=False)
    ]

st.dataframe(
    search_df.style.apply(color_df, axis=1),
    use_container_width=True
)

# =================================================
# 📥 엑셀 다운로드
# =================================================
st.divider()
st.subheader("📥 엑셀 다운로드 (색상 포함)")

if st.button("📥 엑셀 파일 다운로드"):
    # [수정] 버튼 클릭 시 로그 출력
    print("[User Action] 엑셀 다운로드 요청이 들어왔습니다. 파일 생성 중...")

    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    red = PatternFill("solid", start_color="FFCCCC")
    yellow = PatternFill("solid", start_color="FFF2CC")

    remain_col = [cell.value for cell in ws[1]].index("남은일수") + 1

    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=remain_col).value
        if val < 0:
            fill = red
        elif val <= 30:
            fill = yellow
        else:
            continue

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

    final_output = io.BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    print("[System Log] 엑셀 파일 생성 완료. 다운로드 준비 끝.")

    st.download_button(
        label="⬇️ 엑셀 파일 저장",
        data=final_output,
        file_name="시약_유통기한_자동관리_결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
